#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
功耗汇总表自动生成脚本 (纯Python计算版，不依赖LibreOffice/Excel)
用法: python power_summary_pipeline.py

依赖: pip install openpyxl pandas
"""

import re
import csv
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ============================================================
# 配置区：每次换数据只需要改这里
# ============================================================

RAW_DATA_PATH = sys.argv[2]          # 原始采集数据文件 (.csv 或 .xlsx)
CONFIG_PATH = sys.argv[1]             # 模块/网络配置表
CONFIG_SHEET = None                     # 配置表要用哪个sheet, None=用第一个
CONFIG_MODULE_COL = "模块"               # 配置表里"模块"列的表头文字
CONFIG_NET_COL = "输入网络"               # 配置表里"网络"列的表头文字(带_P后缀)

OUTPUT_FILE_NAME = "功耗汇总表.xlsx"

TEST_SHEET_NAME = "测试数据"
CONV_SHEET_NAME = "功耗折算表"
SUMMARY_SHEET_NAME = "功耗汇总表"

# Sub-Module -> Module大类映射，没写到的一律归入 Others，不要瞎猜
MODULE_MAP = {
    "CPU": "1_BB",
    "DDR": "2_MEMORY", "UFS": "2_MEMORY", "SD": "2_MEMORY",
    "OLED": "3_LCM",
    "MAIN CAM": "5_CAMERA", "FRONT CAM": "5_CAMERA", "WIDE 8M": "5_CAMERA", "FLASH": "5_CAMERA",
    "WCN": "8_WIFI_NFC_BT", "NFC": "8_WIFI_NFC_BT",
    "SIM": "10_SIM_T_MOTOR", "TP": "10_SIM_T_MOTOR", "VIB": "10_SIM_T_MOTOR",
}
MODULE_ORDER = ["1_BB", "2_MEMORY", "3_LCM", "5_CAMERA", "8_WIFI_NFC_BT", "10_SIM_T_MOTOR", "Others"]
MODULE_LABELS = {
    "1_BB": "CPU", "2_MEMORY": "DDR", "3_LCM": "屏幕",
    "5_CAMERA": "camera", "8_WIFI_NFC_BT": "WCN", "10_SIM_T_MOTOR": "SIM",
}

# 单个Sub-Module电流超过这个阈值(mA)就提示可能是异常数据，不代表一定错，仅供人工复核
ANOMALY_THRESHOLD_MA = 1000

# ============================================================
# 第一步：读原始数据，自动定位表头行，算出每列平均值
# ============================================================

HEADER_SUFFIX_RE = re.compile(r".+_(I|P|V)$")


def _read_raw_rows(path):
    """读取原始文件所有行，csv/xlsx都支持，返回 list[list]"""
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [["" if v is None else v for v in row] for row in ws.iter_rows(values_only=True)]


def _find_header_row(rows):
    """按 XXX_I / XXX_P / XXX_V 后缀特征找表头行，不依赖固定行号"""
    best_row, best_score = None, 0
    for i, row in enumerate(rows):
        score = sum(1 for cell in row if isinstance(cell, str) and HEADER_SUFFIX_RE.match(cell))
        if score > best_score:
            best_row, best_score = i, score
    if best_row is None:
        raise ValueError("没能在原始数据里找到形如 XXX_I/XXX_P/XXX_V 的表头行，请检查文件")
    return best_row


def _find_first_data_row(rows, header_idx):
    """表头行之后，找第一行第2列是数字的行，认定为数据起始行"""
    for i in range(header_idx + 1, len(rows)):
        val = rows[i][1] if len(rows[i]) > 1 else None
        try:
            float(val)
            return i
        except (TypeError, ValueError):
            continue
    raise ValueError("表头行之后没有找到数值数据，请检查文件")


def load_raw_averages(raw_path):
    """返回 (header列表, {列名: 平均值} 字典, 原始数据DataFrame)"""
    rows = _read_raw_rows(Path(raw_path))
    header_idx = _find_header_row(rows)
    data_start_idx = _find_first_data_row(rows, header_idx)

    header = rows[header_idx]
    df = pd.DataFrame(rows[data_start_idx:], columns=header)

    # 去掉重复/空列名 (比如 System Messages 后面那个没名字的列)
    df = df.loc[:, [c for c in df.columns if c not in ("", None)]]
    df = df.apply(pd.to_numeric, errors="coerce")  # 非数字(如空字符串)转NaN

    averages = df.mean(skipna=True).to_dict()  # NaN会被skipna跳过，等价于之前的IFERROR兜底

    print(f"[原始数据] 表头行=文件第{header_idx+1}行，数据从第{data_start_idx+1}行开始，共{len(df)}行样本")
    return header, averages, df


# ============================================================
# 第二步：按配置表 + 平均值字典，算出每个Sub-Module的Vavg/Iavg/Pavg
# ============================================================

def build_conversion_table(config_path, averages):
    cfg_wb = load_workbook(config_path, data_only=True)
    cfg = cfg_wb[CONFIG_SHEET] if CONFIG_SHEET else cfg_wb[cfg_wb.sheetnames[0]]

    mod_col = net_col = None
    for c in range(1, cfg.max_column + 1):
        v = cfg.cell(row=1, column=c).value
        if v == CONFIG_MODULE_COL and mod_col is None:
            mod_col = c
        elif v == CONFIG_NET_COL and net_col is None:
            net_col = c
    if mod_col is None or net_col is None:
        raise ValueError(f"配置表里没找到列 '{CONFIG_MODULE_COL}' 或 '{CONFIG_NET_COL}'，请检查表头文字是否一致")

    rows_out = []
    for r in range(2, cfg.max_row + 1):
        mod = cfg.cell(row=r, column=mod_col).value
        net = cfg.cell(row=r, column=net_col).value
        if mod is None and net is None:
            continue
        mod = mod.strip() if isinstance(mod, str) else mod
        net = net.strip() if isinstance(net, str) else net
        base = net[:-2] if net and net.endswith("_P") else net

        v_avg = averages.get(f"{base}_V", 0) or 0
        i_avg = averages.get(f"{base}_I", 0) or 0
        p_avg = averages.get(f"{base}_P", 0) or 0
        rows_out.append({
            "Sub-Module": mod, "Net": base,
            "Vavg(V)": v_avg / 1000, "Iavg(mA)": i_avg, "Pavg(mW)": p_avg,
        })

    conv_df = pd.DataFrame(rows_out)
    print(f"[功耗折算表] 共 {len(conv_df)} 行")
    return conv_df


# ============================================================
# 第三步：按 MODULE_MAP 归类汇总，Others兜底
# ============================================================

def build_summary_table(conv_df):
    sub_to_module = conv_df["Sub-Module"].map(MODULE_MAP).fillna("Others")
    grouped = conv_df.groupby(sub_to_module)["Iavg(mA)"].sum()

    total = conv_df["Iavg(mA)"].sum()
    rows_out = [{"Module": "0_BAT", "4V Bat Current(mA)": total, "Label": "电池"}]

    categorized_sum = 0.0
    for mod in MODULE_ORDER:
        if mod == "Others":
            continue
        val = grouped.get(mod, 0.0)
        categorized_sum += val
        rows_out.append({"Module": mod, "4V Bat Current(mA)": val, "Label": MODULE_LABELS.get(mod, "")})

    others_val = total - categorized_sum  # Others = 总合计 - 已归类项，不手动枚举剩余Sub-Module
    rows_out.append({"Module": "Others", "4V Bat Current(mA)": others_val, "Label": ""})

    summary_df = pd.DataFrame(rows_out)
    print(f"[功耗汇总表] 0_BAT={total:.2f}mA, 分类相加={categorized_sum + others_val:.2f}mA "
          f"({'一致 ✓' if abs(total - (categorized_sum + others_val)) < 1e-6 else '不一致 ✗'})")
    return summary_df


# ============================================================
# 第四步：异常值检查
# ============================================================

def check_anomalies(conv_df):
    flagged = conv_df[conv_df["Iavg(mA)"].abs() >= ANOMALY_THRESHOLD_MA]
    if not flagged.empty:
        print(f"[警告] 以下 {len(flagged)} 项电流超过 {ANOMALY_THRESHOLD_MA}mA 阈值，"
              f"疑似原始数据异常，建议人工复核，不要直接采信：")
        for _, row in flagged.iterrows():
            print(f"  - {row['Sub-Module']} ({row['Net']}): {row['Iavg(mA)']:.2f} mA")


# ============================================================
# 写出xlsx
# ============================================================

def write_output(header, df_raw, conv_df, summary_df, output_file):
    wb = Workbook()

    ws = wb.active
    ws.title = TEST_SHEET_NAME
    ws.append(header)
    for row in df_raw.itertuples(index=False):
        ws.append(list(row))

    conv_ws = wb.create_sheet(CONV_SHEET_NAME)
    conv_ws.append(list(conv_df.columns))
    for row in conv_df.itertuples(index=False):
        conv_ws.append(list(row))

    summary_ws = wb.create_sheet(SUMMARY_SHEET_NAME)
    summary_ws.append(list(summary_df.columns))
    for row in summary_df.itertuples(index=False):
        summary_ws.append(list(row))

    header_font = Font(name="Arial", bold=True, size=10)
    body_font = Font(name="Arial", size=10)
    for ws_ in (conv_ws, summary_ws):
        for c in range(1, ws_.max_column + 1):
            ws_.cell(row=1, column=c).font = header_font
        for r in range(2, ws_.max_row + 1):
            for c in range(1, ws_.max_column + 1):
                ws_.cell(row=r, column=c).font = body_font

    wb.save(output_file)
    print(f"[保存] {output_file}")
    os.startfile(output_file)


# ============================================================
# 主流程
# ============================================================

def main():
    header, averages, df_raw = load_raw_averages(RAW_DATA_PATH)
    conv_df = build_conversion_table(CONFIG_PATH, averages)
    summary_df = build_summary_table(conv_df)
    check_anomalies(conv_df)
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = os.path.dirname(SCRIPT_DIR)
    OUTPUT_FILE = os.path.join(BASE_DIR, "OUT", "power", OUTPUT_FILE_NAME)
    print("汇总文件：", OUTPUT_FILE)
    write_output(header, df_raw, conv_df, summary_df, OUTPUT_FILE)

if __name__ == "__main__":
    main()